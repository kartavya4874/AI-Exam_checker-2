"""
Excel report builder for exam results.

Produces multi-sheet Excel workbooks with summary, question-wise breakdown,
individual feedback, and statistics sheets.
"""

import io
import logging
from datetime import datetime
from typing import Dict, List, Any, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# ── Styling constants ─────────────────────────────────────
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
PASS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FAIL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _style_header_row(ws, num_cols: int) -> None:
    """Apply header styling to the first row."""
    for col_idx in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER


def _auto_width(ws) -> None:
    """Auto-fit column widths based on content."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val = str(cell.value) if cell.value else ""
                max_len = max(max_len, len(val))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 3, 50)


def generate_excel_report(
    course_code: str,
    results: List[Dict[str, Any]],
    stats: Optional[Dict[str, Any]] = None,
) -> io.BytesIO:
    """
    Generate a full Excel report for a course.

    Args:
        course_code: The course code (e.g. "CS101").
        results: List of evaluation result dicts, each containing:
            - roll_number, obtained_marks, total_marks, percentage, grade,
              overall_feedback, question_results (list of per-question dicts)
        stats: Optional pre-computed statistics dict with:
            - average, max, min, pass_rate, grade_distribution

    Returns:
        BytesIO buffer containing the Excel workbook.
    """
    wb = Workbook()

    # ── Sheet 1: Summary ──────────────────────────────────
    ws_summary = wb.active
    ws_summary.title = "Summary"
    headers = ["Roll Number", "Obtained Marks", "Total Marks", "Percentage", "Grade"]
    ws_summary.append(headers)
    _style_header_row(ws_summary, len(headers))

    for r in results:
        row_data = [
            r.get("roll_number", ""),
            r.get("obtained_marks", 0),
            r.get("total_marks", 0),
            round(r.get("percentage", 0), 2),
            r.get("grade", ""),
        ]
        ws_summary.append(row_data)
        row_idx = ws_summary.max_row
        grade = r.get("grade", "")
        if grade in ("F",):
            for col_idx in range(1, len(headers) + 1):
                ws_summary.cell(row=row_idx, column=col_idx).fill = FAIL_FILL
        elif grade in ("A+", "A"):
            for col_idx in range(1, len(headers) + 1):
                ws_summary.cell(row=row_idx, column=col_idx).fill = PASS_FILL

    _auto_width(ws_summary)

    # ── Sheet 2: Question-wise Breakdown ──────────────────
    ws_questions = wb.create_sheet("Question Breakdown")

    # Collect all unique question numbers
    all_qnums = set()
    for r in results:
        for qr in r.get("question_results", []):
            all_qnums.add(qr.get("question_number", ""))
    sorted_qnums = sorted(all_qnums, key=_question_sort_key)

    q_headers = ["Roll Number"] + [f"Q{q}" for q in sorted_qnums]
    ws_questions.append(q_headers)
    _style_header_row(ws_questions, len(q_headers))

    for r in results:
        qr_map = {
            qr.get("question_number", ""): qr.get("marks_obtained", 0)
            for qr in r.get("question_results", [])
        }
        row = [r.get("roll_number", "")]
        for qn in sorted_qnums:
            row.append(qr_map.get(qn, 0))
        ws_questions.append(row)

    _auto_width(ws_questions)

    # ── Sheet 3: Individual Feedback ──────────────────────
    ws_feedback = wb.create_sheet("Feedback")
    fb_headers = [
        "Roll Number",
        "Question",
        "Status",
        "Content Type",
        "Marks",
        "Feedback",
        "Error Analysis",
    ]
    ws_feedback.append(fb_headers)
    _style_header_row(ws_feedback, len(fb_headers))

    for r in results:
        for qr in r.get("question_results", []):
            ws_feedback.append(
                [
                    r.get("roll_number", ""),
                    qr.get("question_number", ""),
                    qr.get("status", ""),
                    qr.get("content_type", ""),
                    f"{qr.get('marks_obtained', 0)}/{qr.get('marks_allocated', 0)}",
                    qr.get("feedback", ""),
                    qr.get("error_analysis", ""),
                ]
            )

    _auto_width(ws_feedback)

    # ── Sheet 4: Statistics ───────────────────────────────
    ws_stats = wb.create_sheet("Statistics")

    if stats is None:
        stats = _compute_stats(results)

    stat_rows = [
        ["Statistic", "Value"],
        ["Course Code", course_code],
        ["Total Students", stats.get("total_students", len(results))],
        ["Average Marks", round(stats.get("average", 0), 2)],
        ["Highest Marks", stats.get("max", 0)],
        ["Lowest Marks", stats.get("min", 0)],
        ["Pass Rate (%)", round(stats.get("pass_rate", 0), 2)],
        ["Generated At", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
    ]
    for row in stat_rows:
        ws_stats.append(row)
    _style_header_row(ws_stats, 2)

    # Grade distribution sub-table
    ws_stats.append([])
    ws_stats.append(["Grade", "Count"])
    grade_row = ws_stats.max_row
    ws_stats.cell(row=grade_row, column=1).font = Font(bold=True)
    ws_stats.cell(row=grade_row, column=2).font = Font(bold=True)

    grade_dist = stats.get("grade_distribution", {})
    for grade, count in sorted(grade_dist.items()):
        ws_stats.append([grade, count])

    # ── Bar chart for grade distribution ──
    if grade_dist:
        chart = BarChart()
        chart.type = "col"
        chart.title = f"{course_code} — Grade Distribution"
        chart.y_axis.title = "Students"
        chart.x_axis.title = "Grade"
        chart.style = 10

        data_start = grade_row + 1
        data_end = grade_row + len(grade_dist)
        data_ref = Reference(ws_stats, min_col=2, min_row=data_start, max_row=data_end)
        cats_ref = Reference(ws_stats, min_col=1, min_row=data_start, max_row=data_end)
        chart.add_data(data_ref, titles_from_data=False)
        chart.set_categories(cats_ref)
        chart.shape = 4
        ws_stats.add_chart(chart, "D2")

    _auto_width(ws_stats)

    # ── Write to buffer ───────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    logger.info("Generated Excel report for %s (%d students)", course_code, len(results))
    return buf


def _compute_stats(results: List[Dict[str, Any]]) -> Dict[str, Any]:
    """Compute basic statistics from results list."""
    if not results:
        return {
            "total_students": 0,
            "average": 0,
            "max": 0,
            "min": 0,
            "pass_rate": 0,
            "grade_distribution": {},
        }

    percentages = [r.get("percentage", 0) for r in results]
    grades = [r.get("grade", "F") for r in results]
    grade_dist: Dict[str, int] = {}
    for g in grades:
        grade_dist[g] = grade_dist.get(g, 0) + 1

    passing = sum(1 for p in percentages if p >= 45)  # D or above

    return {
        "total_students": len(results),
        "average": sum(percentages) / len(percentages),
        "max": max(percentages),
        "min": min(percentages),
        "pass_rate": (passing / len(results)) * 100,
        "grade_distribution": grade_dist,
    }


def _question_sort_key(q: str) -> tuple:
    """Sort key for question numbers: numeric first, then alpha sub-parts."""
    import re

    parts = re.findall(r"(\d+|[a-zA-Z]+)", str(q))
    result = []
    for p in parts:
        if p.isdigit():
            result.append((0, int(p), ""))
        else:
            result.append((1, 0, p.lower()))
    return tuple(result) if result else ((0, 0, str(q)),)
